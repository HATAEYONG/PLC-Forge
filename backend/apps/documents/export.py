"""Excel Export (PRD MVP 26번).

설계 산출물을 .xlsx 워크북으로 생성한다. 시트: I/O List, Alarm, Interlock,
Cause & Effect, FAT, SAT. openpyxl 사용, 한글 셀은 UTF-8로 안전하게 기록된다.
"""

import io

from openpyxl import Workbook
from openpyxl.styles import Font

from apps.interlocks.services import cause_effect_matrix

HEADER_FONT = Font(bold=True)


def build_workbook(project) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    # ── I/O List ──
    ws = wb.create_sheet("I_O List")
    ws.append(["Tag", "Signal", "Description", "Source", "SourceRef"])
    for cell in ws[1]:
        cell.font = HEADER_FONT
    for p in project.io_points.all():
        ws.append([p.tag, p.signal_type, p.description, p.source_type, p.source_ref])

    # ── Alarm ──
    ws = wb.create_sheet("Alarm")
    ws.append(["Code", "Source", "Condition", "Priority", "Reset", "Latching", "Message"])
    for cell in ws[1]:
        cell.font = HEADER_FONT
    for a in project.alarms.all():
        ws.append([a.code, a.source, a.condition, a.priority, a.reset_type, a.latching, a.message])

    # ── Interlock ──
    ws = wb.create_sheet("Interlock")
    ws.append(
        [
            "Code",
            "Protected",
            "Condition",
            "Effect",
            "Reset",
            "Safety",
            "BypassPermission",
            "Reason",
        ]
    )
    for cell in ws[1]:
        cell.font = HEADER_FONT
    for i in project.interlocks.all():
        ws.append(
            [
                i.code,
                i.protected_device,
                i.condition,
                i.effect,
                i.reset_condition,
                i.safety_related,
                i.bypass_permission,
                i.reason,
            ]
        )

    # ── Cause & Effect ──
    ce = cause_effect_matrix(project=project)
    ws = wb.create_sheet("Cause_Effect")
    ws.append(["Cause \\ Effect", *ce["effects"]])
    for cell in ws[1]:
        cell.font = HEADER_FONT
    for row in ce["matrix"]:
        ws.append([row["cause"], *["X" if row["effects"].get(e) else "" for e in ce["effects"]]])

    # ── FAT / SAT ──
    for phase in ["FAT", "SAT"]:
        ws = wb.create_sheet(phase)
        ws.append(
            ["TestID", "Category", "Precondition", "Procedure", "Expected", "Status", "Source"]
        )
        for cell in ws[1]:
            cell.font = HEADER_FONT
        for t in project.test_cases.filter(phase=phase):
            ws.append(
                [
                    t.test_id,
                    t.category,
                    t.precondition,
                    t.procedure,
                    t.expected_result,
                    t.status,
                    f"{t.source_type}:{t.source_ref}",
                ]
            )

    return wb


def export_project_xlsx(project) -> bytes:
    wb = build_workbook(project)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
