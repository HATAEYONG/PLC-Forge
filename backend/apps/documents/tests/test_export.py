"""Excel Export 테스트 (PRD MVP 26번)."""

import io

import pytest
from openpyxl import load_workbook

from apps.alarms.services import generate_alarms
from apps.design.models import RiskLevel
from apps.design.services import create_design_decision
from apps.documents.export import build_workbook
from apps.interlocks.services import generate_interlocks
from apps.io_points.services import estimate_io
from apps.projects.tests.factories import ProjectFactory
from apps.requirements.models import SourceType, ValueType
from apps.requirements.services import create_fact


@pytest.fixture
def project(db):
    p = ProjectFactory(code="EXP-1")
    fact = create_fact(
        project=p,
        fact_key="DEVICE_LIST",
        value_json=["펌프"],
        value_type=ValueType.LIST,
        source_type=SourceType.MANUAL,
    )
    estimate_io(project=p)
    create_design_decision(
        project=p,
        decision_type="ALARM_REQUIREMENT",
        subject_type="ALARM",
        subject_id="TEMP_HIGH",
        decision_value={"type": "HIGH_TEMP", "priority": "HIGH"},
        reason="과열 알람",
        input_facts=[fact],
    )
    create_design_decision(
        project=p,
        decision_type="INTERLOCK_REQUIREMENT",
        subject_type="INTERLOCK",
        subject_id="ESTOP",
        decision_value={"effect": "STOP"},
        reason="비상정지",
        input_facts=[fact],
        risk_level=RiskLevel.CRITICAL,
    )
    generate_alarms(project=p)
    generate_interlocks(project=p)
    return p


@pytest.mark.django_db
def test_workbook_has_expected_sheets(project):
    wb = build_workbook(project)
    assert set(wb.sheetnames) == {"I_O List", "Alarm", "Interlock", "Cause_Effect", "FAT", "SAT"}


@pytest.mark.django_db
def test_io_sheet_contains_points(project):
    wb = build_workbook(project)
    ws = wb["I_O List"]
    # 헤더 + 펌프 I/O 3점(DO+2DI)
    assert ws.max_row == 1 + project.io_points.count()
    assert ws["A1"].value == "Tag"


@pytest.mark.django_db
def test_korean_cells_roundtrip(project):
    wb = build_workbook(project)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    reopened = load_workbook(buffer)
    ws = reopened["Interlock"]
    # 인터록 condition/effect 등 한글이 깨지지 않고 보존됨
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert rows  # 최소 1건
    assert any("비상정지" in str(cell) for row in rows for cell in row if cell)


@pytest.mark.django_db
def test_export_api_returns_xlsx(api_client, project):
    response = api_client.get(f"/api/projects/{project.id}/export/")
    assert response.status_code == 200
    assert "spreadsheetml" in response["Content-Type"]
    assert f"PLC-Forge_{project.code}.xlsx" in response["Content-Disposition"]
    # 실제 xlsx 바이트로 로드 가능
    wb = load_workbook(io.BytesIO(response.content))
    assert "FAT" in wb.sheetnames
